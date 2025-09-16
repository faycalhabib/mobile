"""
Test simple pour vérifier la copie et modification du template
"""
import shutil
import os
from openpyxl import load_workbook

# Chemins
template_path = r"C:\Users\faycalhabibahmat\Desktop\Moov\UGP\Rapport UGP.xlsx"
output_path = r"C:\Users\faycalhabibahmat\Desktop\Moov\UGP_Reporter\outputs\test_simple.xlsx"

try:
    print("1. Copie du template...")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy2(template_path, output_path)
    print("   ✓ Template copié")
    
    print("2. Ouverture avec openpyxl...")
    wb = load_workbook(output_path)
    print(f"   ✓ Ouvert, feuilles: {wb.sheetnames}")
    
    print("3. Modification simple...")
    ws = wb.active
    # Essayer d'écrire dans une cellule simple
    ws['A20'] = "TEST"
    print("   ✓ Cellule modifiée")
    
    print("4. Sauvegarde...")
    wb.save(output_path)
    wb.close()
    print("   ✓ Fichier sauvegardé")
    
    print(f"\n✅ Test réussi! Fichier: {output_path}")
    
except Exception as e:
    print(f"\n❌ Erreur: {e}")
    import traceback
    traceback.print_exc()
