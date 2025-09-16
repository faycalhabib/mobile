"""
Module pour remplir le template Excel existant sans modifier sa structure
Préserve les logos, signatures, et mise en forme
"""
import os
import shutil
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as OpenpyxlImage
import logging

logger = logging.getLogger(__name__)


class TemplateFiller:
    """Remplit le template Excel existant avec les nouvelles données"""
    
    def __init__(self):
        self.template_path = r"C:\Users\faycalhabibahmat\Desktop\Moov\UGP\Rapport UGP.xlsx"
        
    def fill_template(self, 
                      processed_df: pd.DataFrame,
                      metadata: dict,
                      output_path: str) -> str:
        """
        Remplit le template existant avec les données sans modifier la structure
        
        Args:
            processed_df: DataFrame avec les données traitées
            metadata: Métadonnées du rapport
            output_path: Chemin de sortie
            
        Returns:
            Chemin du fichier généré
        """
        try:
            # Vérifier que le template existe
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template non trouvé: {self.template_path}")
            
            # Créer le dossier de sortie
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Copier le template vers le fichier de sortie
            # IMPORTANT: Utiliser copy2 pour préserver les métadonnées
            shutil.copy2(self.template_path, output_path)
            logger.info(f"Template copié vers: {output_path}")
            
            # Charger le workbook avec openpyxl
            # Charger simplement sans options complexes pour éviter la corruption
            wb = load_workbook(output_path)
            
            # Obtenir la première feuille (ou la feuille active)
            # Le template a une feuille nommée 'Rapport paiement'
            if 'Rapport paiement' in wb.sheetnames:
                ws = wb['Rapport paiement']
            elif 'Rapport' in wb.sheetnames:
                ws = wb['Rapport']
            else:
                ws = wb.active
                logger.info(f"Utilisation de la feuille active: {ws.title}")
            
            # Remplir les métadonnées dans les cellules spécifiques
            # En analysant le template, on remplit les bonnes cellules
            self._fill_metadata(ws, metadata)
            
            # Ajouter le logo
            self._add_logo(ws)
            
            # Remplir les données des transactions
            self._fill_transactions(ws, processed_df)
            
            # Sauvegarder le fichier
            wb.save(output_path)
            wb.close()
            logger.info("Fichier sauvegardé avec succès")
            
            logger.info(f"Template rempli avec succès: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Erreur lors du remplissage du template: {e}")
            raise
    
    def _write_to_cell(self, worksheet, row, col, value):
        """Écrit dans une cellule en gérant les cellules fusionnées"""
        cell = worksheet.cell(row=row, column=col)
        
        # Si c'est une cellule fusionnée, trouver la cellule principale
        if isinstance(cell, MergedCell):
            # Trouver la plage fusionnée qui contient cette cellule
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Écrire dans la cellule en haut à gauche de la plage fusionnée
                    main_cell = worksheet.cell(row=merged_range.min_row, 
                                              column=merged_range.min_col)
                    if main_cell.value is None or main_cell.value == '':
                        main_cell.value = value
                    return True
        else:
            # Cellule normale
            if cell.value is None or cell.value == '':
                cell.value = value
            return True
        return False
    
    def _fill_metadata(self, worksheet, metadata: dict):
        """
        Remplit les métadonnées dans les cellules appropriées
        Recherche les cellules par leur contenu pour être plus flexible
        """
        # Parcourir les premières lignes pour trouver les emplacements
        for row in range(1, 15):  # Chercher dans les 15 premières lignes
            for col in range(1, 10):  # Chercher dans les 10 premières colonnes
                cell = worksheet.cell(row=row, column=col)
                
                # Gérer les cellules fusionnées pour la lecture
                cell_value = ''
                if isinstance(cell, MergedCell):
                    # Pour les cellules fusionnées, chercher la valeur dans la cellule principale
                    for merged_range in worksheet.merged_cells.ranges:
                        if merged_range.min_row <= row <= merged_range.max_row and \
                           merged_range.min_col <= col <= merged_range.max_col:
                            main_cell = worksheet.cell(row=merged_range.min_row,
                                                      column=merged_range.min_col)
                            cell_value = str(main_cell.value or '').lower().strip()
                            break
                else:
                    cell_value = str(cell.value or '').lower().strip()
                
                # Date de paiement
                if 'date de paiement' in cell_value:
                    # La valeur est généralement dans la cellule suivante
                    value = metadata.get('date_paiement', datetime.now().strftime("%d-%b-%Y"))
                    if self._write_to_cell(worksheet, row, col+1, value):
                        logger.info(f"Date remplie en cellule {get_column_letter(col+1)}{row}")
                
                # Libellé de l'opération
                elif 'libellé' in cell_value or 'libelle' in cell_value:
                    value = metadata.get('libelle', 'PAIEMENT')
                    if self._write_to_cell(worksheet, row, col+1, value):
                        logger.info(f"Libellé rempli en cellule {get_column_letter(col+1)}{row}")
                
                # Budget
                elif 'budget' in cell_value:
                    budget_value = metadata.get('budget', 500000)
                    # Formatter le budget avec séparateurs de milliers
                    value = f"{budget_value:,.0f}".replace(',', ' ')
                    if self._write_to_cell(worksheet, row, col+1, value):
                        logger.info(f"Budget rempli en cellule {get_column_letter(col+1)}{row}")
                
                # Projet
                elif 'projet' in cell_value:
                    value = metadata.get('projet', 'UGP')
                    if self._write_to_cell(worksheet, row, col+1, value):
                        logger.info(f"Projet rempli en cellule {get_column_letter(col+1)}{row}")
    
    def _add_logo(self, worksheet):
        """Vérifier et préserver les images existantes dans le template"""
        try:
            # NE PAS ajouter de nouveau logo si le template en contient déjà un
            # Les images du template original sont automatiquement préservées
            # par la copie du fichier et le chargement avec keep_links=True
            
            # Compter les images existantes
            existing_images = len(worksheet._images) if hasattr(worksheet, '_images') else 0
            logger.info(f"Images existantes dans le template: {existing_images}")
            
            # Si pas d'images dans le template, on peut en ajouter une
            if existing_images == 0:
                logo_path = r"C:\Users\faycalhabibahmat\Desktop\Moov\UGP_Reporter\image\logo.png"
                
                if os.path.exists(logo_path):
                    img = OpenpyxlImage(logo_path)
                    img.width = 150
                    img.height = 60
                    worksheet.add_image(img, 'E1')
                    logger.info("Logo ajouté car aucune image dans le template")
                else:
                    logger.warning(f"Logo non trouvé: {logo_path}")
            else:
                logger.info("Images du template original préservées")
            
        except Exception as e:
            logger.warning(f"Info logo: {e}")
    
    def _fill_transactions(self, worksheet, df: pd.DataFrame):
        """
        Remplit les données de transactions dans le tableau
        Trouve automatiquement où commencer en cherchant les en-têtes
        """
        # Chercher la ligne d'en-tête du tableau
        header_row = None
        header_markers = ['Date', 'N° Transaction', 'Type', 'Statut', 'Montant']
        
        for row in range(1, 30):  # Chercher dans les 30 premières lignes
            row_values = []
            for col in range(1, 15):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            # Vérifier si cette ligne contient les marqueurs d'en-tête
            matches = sum(1 for marker in header_markers if any(marker in val for val in row_values))
            if matches >= 3:  # Au moins 3 marqueurs trouvés
                header_row = row
                logger.info(f"En-tête du tableau trouvé à la ligne {header_row}")
                break
        
        if not header_row:
            # Utiliser une valeur par défaut
            header_row = 8
            logger.warning(f"En-tête non trouvé, utilisation de la ligne {header_row} par défaut")
        
        # Mapper les colonnes
        column_mapping = self._get_column_mapping(worksheet, header_row)
        
        # Commencer à remplir à partir de la ligne suivante
        start_row = header_row + 1
        
        # Remplir les données
        for idx, record in df.iterrows():
            current_row = start_row + idx
            
            # Date
            if 'Date' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Date'], record.get('Date', ''))
            
            # N° Transaction
            if 'Transaction' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Transaction'], record.get('TransactionID', ''))
            
            # Type
            if 'Type' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Type'], record.get('Type', 'PAIEMENT'))
            
            # Statut
            if 'Statut' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Statut'], record.get('Status', ''))
            
            # Montant
            if 'Montant' in column_mapping:
                amount = record.get('Amount', 0)
                # Formatter avec séparateurs de milliers
                self._write_to_cell(worksheet, current_row, column_mapping['Montant'], f"{amount:,.0f}".replace(',', ' '))
            
            # Frais
            if 'Frais' in column_mapping:
                fee = record.get('Frais', 0)
                self._write_to_cell(worksheet, current_row, column_mapping['Frais'], f"{fee:,.0f}".replace(',', ' '))
            
            # De
            if 'De' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['De'], record.get('De', 'UGP'))
            
            # Vers
            if 'Vers' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Vers'], record.get('Vers', ''))
            
            # Bénéficiaire
            if 'Beneficiaire' in column_mapping:
                self._write_to_cell(worksheet, current_row, column_mapping['Beneficiaire'], record.get('Beneficiaire', ''))
        
        # Ajouter la ligne de total
        if len(df) > 0:
            total_row = start_row + len(df) + 1
            
            # Écrire "TOTAL" dans la colonne Statut ou Type
            if 'Statut' in column_mapping:
                self._write_to_cell(worksheet, total_row, column_mapping['Statut'], "TOTAL:")
            elif 'Type' in column_mapping:
                self._write_to_cell(worksheet, total_row, column_mapping['Type'], "TOTAL:")
            
            # Total des montants
            if 'Montant' in column_mapping:
                total_amount = df['Amount'].sum()
                self._write_to_cell(worksheet, total_row, column_mapping['Montant'], f"{total_amount:,.0f}".replace(',', ' '))
            
            # Total des frais
            if 'Frais' in column_mapping:
                total_fees = df['Frais'].sum()
                self._write_to_cell(worksheet, total_row, column_mapping['Frais'], f"{total_fees:,.0f}".replace(',', ' '))
            
            logger.info(f"Totaux ajoutés à la ligne {total_row}")
    
    def _get_column_mapping(self, worksheet, header_row: int) -> dict:
        """
        Crée un mapping des noms de colonnes vers leurs indices
        """
        mapping = {}
        
        for col in range(1, 20):  # Vérifier les 20 premières colonnes
            cell = worksheet.cell(row=header_row, column=col)
            
            # Gérer les cellules fusionnées
            if isinstance(cell, MergedCell):
                # Pour les cellules fusionnées, chercher la valeur dans la cellule principale
                for merged_range in worksheet.merged_cells.ranges:
                    if merged_range.min_row <= header_row <= merged_range.max_row and \
                       merged_range.min_col <= col <= merged_range.max_col:
                        main_cell = worksheet.cell(row=merged_range.min_row,
                                                  column=merged_range.min_col)
                        cell_value = main_cell.value
                        break
                else:
                    cell_value = None
            else:
                cell_value = cell.value
            
            if cell_value:
                value_lower = str(cell_value).lower().strip()
                
                if 'date' in value_lower:
                    mapping['Date'] = col
                elif 'transaction' in value_lower or 'n°' in value_lower:
                    mapping['Transaction'] = col
                elif 'type' in value_lower:
                    mapping['Type'] = col
                elif 'statut' in value_lower or 'status' in value_lower:
                    mapping['Statut'] = col
                elif 'montant' in value_lower and 'frais' not in value_lower:
                    mapping['Montant'] = col
                elif 'frais' in value_lower:
                    mapping['Frais'] = col
                elif 'de' in value_lower and len(value_lower) < 5:
                    mapping['De'] = col
                elif 'vers' in value_lower or 'à' in value_lower:
                    mapping['Vers'] = col
                elif 'bénéficiaire' in value_lower or 'beneficiaire' in value_lower:
                    mapping['Beneficiaire'] = col
        
        logger.info(f"Mapping des colonnes: {mapping}")
        return mapping
