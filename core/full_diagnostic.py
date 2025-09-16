"""
Système de diagnostic complet pour tracer TOUT le processus
De la lecture à l'écriture finale dans Excel
"""
import pandas as pd
import logging
import os
import json
from datetime import datetime
import win32com.client as win32
import pythoncom

logger = logging.getLogger(__name__)


class FullDiagnostic:
    """Scanner complet pour identifier le problème"""
    
    def __init__(self):
        self.diagnostic_results = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'steps': [],
            'errors': [],
            'data_flow': {},
            'final_status': 'UNKNOWN'
        }
        
        # Configuration des logs avec plus de détails
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - [%(levelname)s] - %(message)s'
        )
    
    def scan_full_process(self, bulk_path, export_path, fees_path, output_path):
        """Scanner tout le processus de bout en bout"""
        print("\n" + "="*80)
        print(" 🔍 DIAGNOSTIC COMPLET DU SYSTÈME")
        print("="*80)
        
        # Étape 1: Vérifier les fichiers
        self._step("VÉRIFICATION DES FICHIERS D'ENTRÉE")
        files_ok = self._check_input_files(bulk_path, export_path, fees_path)
        if not files_ok:
            self._error("Fichiers d'entrée non valides")
            return self.diagnostic_results
        
        # Étape 2: Lire BulkReport
        self._step("LECTURE DU BULKREPORT")
        bulk_df = self._read_bulk_detailed(bulk_path)
        
        # Étape 3: Lire Export
        self._step("LECTURE DU FICHIER EXPORT")
        export_df = self._read_export_detailed(export_path)
        
        # Étape 4: Traiter les données
        self._step("TRAITEMENT DES DONNÉES")
        processed_df = self._process_data_detailed(bulk_df, export_df)
        
        # Étape 5: Vérifier le template
        self._step("VÉRIFICATION DU TEMPLATE EXCEL")
        template_ok = self._check_template()
        
        # Étape 6: Écriture dans Excel
        self._step("ÉCRITURE DANS EXCEL")
        write_ok = self._test_excel_write(processed_df, output_path)
        
        # Étape 7: Vérification finale
        self._step("VÉRIFICATION DU FICHIER FINAL")
        final_ok = self._verify_final_file(output_path)
        
        # Résumé
        self._print_summary()
        
        # Sauvegarder le diagnostic
        self._save_diagnostic()
        
        return self.diagnostic_results
    
    def _step(self, title):
        """Marquer une nouvelle étape"""
        print(f"\n{'='*60}")
        print(f" 📌 {title}")
        print(f"{'='*60}")
        self.diagnostic_results['steps'].append({
            'title': title,
            'timestamp': datetime.now().isoformat(),
            'status': 'STARTED'
        })
    
    def _check_input_files(self, bulk_path, export_path, fees_path):
        """Vérifier l'existence et la lisibilité des fichiers"""
        files_ok = True
        
        for name, path in [('BulkReport', bulk_path), ('Export', export_path), ('Fees', fees_path)]:
            if path and os.path.exists(path):
                size = os.path.getsize(path)
                print(f"  ✓ {name}: {os.path.basename(path)} ({size} bytes)")
                self.diagnostic_results['data_flow'][f'{name}_path'] = path
                self.diagnostic_results['data_flow'][f'{name}_size'] = size
            else:
                print(f"  ✗ {name}: Fichier non trouvé - {path}")
                files_ok = False
                self._error(f"Fichier {name} non trouvé: {path}")
        
        return files_ok
    
    def _read_bulk_detailed(self, bulk_path):
        """Lire BulkReport avec diagnostic détaillé"""
        print("\n  Lecture du BulkReport:")
        
        try:
            from core.file_handler import FileHandler
            handler = FileHandler()
            
            # Lire avec le handler
            bulk_df, metadata = handler.read_bulk_report(bulk_path)
            
            print(f"    • Lignes lues: {len(bulk_df)}")
            print(f"    • Colonnes: {list(bulk_df.columns[:5])}...")
            
            if len(bulk_df) > 0:
                print(f"\n    Première transaction:")
                for col in ['Credit Msisdn', 'Amount', 'TransactionID']:
                    if col in bulk_df.columns:
                        print(f"      - {col}: {bulk_df.iloc[0][col]}")
            
            self.diagnostic_results['data_flow']['bulk_rows'] = len(bulk_df)
            self.diagnostic_results['data_flow']['bulk_columns'] = list(bulk_df.columns)
            
            # Vérifier pourquoi 0 lignes
            if len(bulk_df) == 0:
                print("\n    ⚠ PROBLÈME: 0 lignes lues!")
                print("    Tentative de lecture directe...")
                
                # Lecture directe pour diagnostic
                with open(bulk_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    print(f"    Fichier contient {len(lines)} lignes au total")
                    
                    # Chercher les lignes avec des transactions
                    for i, line in enumerate(lines):
                        if '23596771275' in line or 'CI9510O2KX' in line:
                            print(f"    Transaction trouvée à ligne {i+1}: {line[:80]}...")
                
            return bulk_df
            
        except Exception as e:
            print(f"    ❌ Erreur: {e}")
            self._error(f"Erreur lecture BulkReport: {e}")
            return pd.DataFrame()
    
    def _read_export_detailed(self, export_path):
        """Lire Export avec diagnostic détaillé"""
        print("\n  Lecture du fichier Export:")
        
        try:
            from core.file_handler import FileHandler
            handler = FileHandler()
            
            export_df = handler.read_export_file(export_path)
            
            print(f"    • Lignes lues: {len(export_df)}")
            print(f"    • Colonnes: {list(export_df.columns)}")
            
            if len(export_df) > 0:
                print(f"\n    Premier bénéficiaire:")
                print(f"      - {export_df.iloc[0].to_dict()}")
            
            self.diagnostic_results['data_flow']['export_rows'] = len(export_df)
            
            return export_df
            
        except Exception as e:
            print(f"    ❌ Erreur: {e}")
            self._error(f"Erreur lecture Export: {e}")
            return pd.DataFrame()
    
    def _process_data_detailed(self, bulk_df, export_df):
        """Traiter les données avec diagnostic"""
        print("\n  Traitement des données:")
        
        try:
            from core.data_processor import DataProcessor
            processor = DataProcessor()
            
            metadata = {
                'date_paiement': '16/09/2025',
                'libelle': 'TEST DIAGNOSTIC',
                'budget': 500000,
                'projet': 'UGP'
            }
            
            processed_df, errors = processor.process_transactions(
                bulk_df, export_df, pd.DataFrame(), metadata
            )
            
            print(f"    • Transactions traitées: {len(processed_df)}")
            
            if len(processed_df) > 0:
                print(f"    • Colonnes finales: {list(processed_df.columns)}")
                print(f"\n    Première transaction traitée:")
                for col in processed_df.columns:
                    print(f"      - {col}: {processed_df.iloc[0][col]}")
            else:
                print("    ⚠ AUCUNE TRANSACTION TRAITÉE!")
            
            if errors:
                print(f"\n    Erreurs/Warnings:")
                for error in errors:
                    print(f"      - {error}")
            
            self.diagnostic_results['data_flow']['processed_rows'] = len(processed_df)
            
            return processed_df
            
        except Exception as e:
            print(f"    ❌ Erreur: {e}")
            self._error(f"Erreur traitement: {e}")
            return pd.DataFrame()
    
    def _check_template(self):
        """Vérifier le template Excel"""
        print("\n  Vérification du template:")
        
        template_path = r"C:\Users\faycalhabibahmat\Desktop\Moov\UGP\Rapport UGP.xlsx"
        
        if os.path.exists(template_path):
            print(f"    ✓ Template trouvé: {os.path.basename(template_path)}")
            
            # Vérifier la feuille
            try:
                import openpyxl
                wb = openpyxl.load_workbook(template_path, read_only=True)
                if 'Rapport paiement' in wb.sheetnames:
                    print(f"    ✓ Feuille 'Rapport paiement' présente")
                    return True
                else:
                    print(f"    ✗ Feuille 'Rapport paiement' manquante")
                    print(f"    Feuilles disponibles: {wb.sheetnames}")
                    return False
            except Exception as e:
                print(f"    ❌ Erreur vérification: {e}")
                return False
        else:
            print(f"    ✗ Template non trouvé: {template_path}")
            return False
    
    def _test_excel_write(self, processed_df, output_path):
        """Tester l'écriture dans Excel avec win32com"""
        print("\n  Test d'écriture Excel:")
        
        if len(processed_df) == 0:
            print("    ⚠ Pas de données à écrire (DataFrame vide)")
            return False
        
        try:
            # Test direct avec win32com
            pythoncom.CoInitialize()
            
            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Ouvrir le fichier de sortie
            wb = excel.Workbooks.Open(os.path.abspath(output_path))
            sheet = wb.Worksheets('Rapport paiement')
            
            print(f"    Écriture de {len(processed_df)} transactions...")
            
            # Écrire directement à la ligne 12
            start_row = 12
            for idx, row in processed_df.iterrows():
                excel_row = start_row + idx
                
                # Écrire les données
                sheet.Cells(excel_row, 2).Value = str(row.get('Date', ''))
                sheet.Cells(excel_row, 3).Value = str(row.get('TransactionID', ''))
                sheet.Cells(excel_row, 4).Value = 'PAIEMENT'
                sheet.Cells(excel_row, 5).Value = str(row.get('Status', 'Success'))
                sheet.Cells(excel_row, 6).Value = str(int(row.get('Amount', 0)))
                sheet.Cells(excel_row, 7).Value = str(int(row.get('Frais', 0)))
                sheet.Cells(excel_row, 8).Value = 'UGP'
                sheet.Cells(excel_row, 9).Value = str(row.get('Vers', ''))
                sheet.Cells(excel_row, 10).Value = str(row.get('Beneficiaire', ''))
                
                print(f"    ✓ Ligne {excel_row} écrite")
            
            # Sauvegarder
            wb.Save()
            print(f"    ✓ Fichier sauvegardé")
            
            # Vérifier que les données sont bien écrites
            print(f"\n    Vérification des données écrites:")
            for i in range(min(2, len(processed_df))):
                excel_row = start_row + i
                value = sheet.Cells(excel_row, 3).Value  # TransactionID
                print(f"      Ligne {excel_row}, col 3: {value}")
            
            wb.Close(False)
            excel.Quit()
            
            pythoncom.CoUninitialize()
            return True
            
        except Exception as e:
            print(f"    ❌ Erreur écriture: {e}")
            self._error(f"Erreur écriture Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _verify_final_file(self, output_path):
        """Vérifier le fichier final"""
        print("\n  Vérification du fichier final:")
        
        if not os.path.exists(output_path):
            print(f"    ✗ Fichier non trouvé: {output_path}")
            return False
        
        try:
            # Lire avec openpyxl pour vérifier
            import openpyxl
            wb = openpyxl.load_workbook(output_path, data_only=True)
            sheet = wb['Rapport paiement']
            
            # Vérifier les données à la ligne 12
            row_12_data = []
            for col in range(2, 11):  # Colonnes B à J
                value = sheet.cell(row=12, column=col).value
                row_12_data.append(value)
            
            print(f"    Données ligne 12: {row_12_data}")
            
            # Vérifier si des données sont présentes
            if any(row_12_data):
                print(f"    ✓ Des données sont présentes")
                return True
            else:
                print(f"    ✗ Aucune donnée trouvée ligne 12")
                return False
                
        except Exception as e:
            print(f"    ❌ Erreur vérification: {e}")
            return False
    
    def _error(self, message):
        """Enregistrer une erreur"""
        self.diagnostic_results['errors'].append({
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
    
    def _print_summary(self):
        """Afficher le résumé du diagnostic"""
        print("\n" + "="*80)
        print(" 📊 RÉSUMÉ DU DIAGNOSTIC")
        print("="*80)
        
        if self.diagnostic_results['errors']:
            print("\n❌ ERREURS DÉTECTÉES:")
            for error in self.diagnostic_results['errors']:
                print(f"  • {error['message']}")
        
        print("\n📈 FLUX DE DONNÉES:")
        for key, value in self.diagnostic_results['data_flow'].items():
            if not key.endswith('_path') and not key.endswith('columns'):
                print(f"  • {key}: {value}")
        
        # Déterminer le statut final
        if self.diagnostic_results['data_flow'].get('processed_rows', 0) > 0:
            self.diagnostic_results['final_status'] = 'DATA_PROCESSED'
        elif self.diagnostic_results['data_flow'].get('bulk_rows', 0) > 0:
            self.diagnostic_results['final_status'] = 'DATA_READ_BUT_NOT_PROCESSED'
        else:
            self.diagnostic_results['final_status'] = 'NO_DATA_READ'
        
        print(f"\n🏁 STATUT FINAL: {self.diagnostic_results['final_status']}")
        
        # Recommandations
        print("\n💡 RECOMMANDATIONS:")
        if self.diagnostic_results['final_status'] == 'NO_DATA_READ':
            print("  1. Vérifier le format du fichier BulkReport")
            print("  2. Vérifier que les données sont bien à la ligne 14-15")
            print("  3. Vérifier les séparateurs (virgules, tabs)")
        elif self.diagnostic_results['final_status'] == 'DATA_READ_BUT_NOT_PROCESSED':
            print("  1. Vérifier le mapping des colonnes")
            print("  2. Vérifier les noms de colonnes dans BulkReport")
        else:
            print("  1. Vérifier que FinalExcelFiller est bien utilisé")
            print("  2. Vérifier les permissions d'écriture")
    
    def _save_diagnostic(self):
        """Sauvegarder le diagnostic dans un fichier"""
        diagnostic_file = f"./logs/diagnostic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        os.makedirs(os.path.dirname(diagnostic_file), exist_ok=True)
        
        with open(diagnostic_file, 'w', encoding='utf-8') as f:
            json.dump(self.diagnostic_results, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n💾 Diagnostic sauvé: {diagnostic_file}")
